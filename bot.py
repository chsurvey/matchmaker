import discord
import openpyxl
import os

def isCodeRight(input_s):
    n_count = 0
    e_count = 0
    for c in input_s:
        if ord('a') <= ord(c.lower()) <= ord('z'):
            e_count+=1
        elif ord('0') <= ord(c.lower()) <= ord('9'):
            n_count+=1
    return True if n_count + e_count == 5 else False

admin_id=283788425264365569

client = discord.Client()
guild = discord.Guild

@client.event
async def on_ready():
    await client.change_presence(activity=discord.Game(name="-도움"))
    print("Ready")

@client.event
async def on_message(message):
    if message.author == client.user:
        return

    elif message.content.startswith('-'):

        cmd = message.content.split()[0].replace("-","")

        if len(message.content.split()) > 1:
            parameters = message.content.split()[1:]

        # if cmd == 'dm':
        #     user = await client.fetch_user(int(parameters[0]))
        #     await user.create_dm()
        #     channel = user.dm_channel
        #     await channel.send("test")
        if cmd == 'mm':
            file = openpyxl.load_workbook("mmlist.xlsx")
            sheet = file.active
            max_row = sheet.max_row
            if(sheet["A1"].value==None):
                max_row = 0
            try:
                print(parameters)
                if message.author.id==admin_id:
                    user = await client.fetch_user(int(parameters[0]))
                    name=user.name
                    for idx in range(1,max_row+1):
                        if sheet["A"+str(idx)].value == str(parameters[0]):
                            sheet.delete_rows(idx)
                            await message.channel.send(name+"님이 mathcmaker에서 제거되었습니다.")
                            file.save("mmlist.xlsx")
                            return

                    sheet["A"+str(max_row+1)].value = str(parameters[0])
                    await message.channel.send(name+"님이 matchmaker에 등록되었습니다.")
                    file.save("mmlist.xlsx")
                    return

                else:
                    await message.channel.send("권한이 없습니다.")
                    return
                    
            except:
                for idx in range(1,max_row+1):
                    if sheet["A"+str(idx)].value == str(message.author.id):
                        sheet.delete_rows(idx)
                        await message.channel.send("mathcmaker에서 제거되었습니다.")
                        file.save("mmlist.xlsx")
                        return

                sheet["A"+str(max_row+1)].value = str(message.author.id)
                await message.channel.send("matchmaker에 등록되었습니다.")
                file.save("mmlist.xlsx")

        if cmd == '도움':
            embed=discord.Embed(title="도움말", color=0x602dd9)
            embed.add_field(name="-방 생성", value="방을 생성합니다", inline=True)
            embed.add_field(name="-방 삭제", value="방을 삭제합니다", inline=True)
            embed.add_field(name="-방 조회", value="현재 있는 방을 나열합니다", inline=True)
            await message.channel.send(embed=embed)

        if cmd == '방':            
        
            if parameters[0] == '삭제':
                file = openpyxl.load_workbook("room.xlsx")
                sheet = file.active
                max_row = sheet.max_row

                if len(parameters)>1 and message.author.id==admin_id:
                    for idx in range(1,max_row+1):
                        if sheet["A"+str(idx)].value==parameters[1]:
                            sheet.delete_rows(idx)
                            file.save("room.xlsx")
                            await message.channel.send("방을 성공적으로 지웠습니다")
                            break
                
                elif len(parameters)>1:
                    await message.channel.send("권한이 없습니다.")

                else:
                    for idx in range(1,max_row+1):
                        if sheet["D"+str(idx)].value==str(message.author.id):
                            sheet.delete_rows(idx)
                            await message.channel.send("방을 성공적으로 지웠습니다")
                            break
                    file.save("room.xlsx")

            if parameters[0] == '조회':
                
                file = openpyxl.load_workbook("room.xlsx")
                sheet = file.active
                max_row = sheet.max_row
                channel = message.channel

                #print(max_row)
                if(sheet["A1"].value==None):
                    max_row = 0
                # sheet[A(n)] = 방장 닉네임
                # sheet[B(n)] = 코드
                # sheet[C(n)] = 비밀번호
                # sheet[D(n)] = user id(식별용)
                # sheet[E(n)] = user name

                if(max_row==0):
                    await channel.send("방이 없습니다!")
                for idx in range(1,max_row+1):
                    
                    host = sheet["A"+str(idx)].value
                    code = sheet["B"+str(idx)].value
                    pw = str(sheet["C"+str(idx)].value)
                    usr_name=sheet["E"+str(idx)].value

                    embed=discord.Embed(title="Room No."+str(idx), color=0x602dd9)
                    embed.add_field(name="Host", value=host+"("+usr_name+")", inline=False)
                    embed.add_field(name="Room Code", value=code, inline=True)
                    embed.add_field(name="Password", value=pw, inline=True)
                      
                    await channel.send(embed=embed)
                
                file.save("room.xlsx")


            if parameters[0] == '생성':
                file = openpyxl.load_workbook("room.xlsx")
                sheet = file.active
                max_row = sheet.max_row
                if(sheet["A1"].value==None):
                    max_row = 0
                #print(max_row)
                
                # sheet[A(n)] = 방장 닉네임
                # sheet[B(n)] = 코드
                # sheet[C(n)] = 비밀번호
                # sheet[D(n)] = user id(식별용)
                # sheet[E(n)] = user name

                if len(parameters) < 4:
                    embed=discord.Embed(title="도움말")
                    embed.add_field(name="방 생성 방법", value="-방 생성 인게임id 방코드 비밀번호", inline=True)
                    
                    await message.channel.send(embed=embed)

                else:
                    #print(parameters)
                    if len(parameters[2]) != 5:
                        await message.channel.send("방 코드는 5자리의 숫자/영문으로 이루어져 있어야 합니다.")
                        return
                    
                    if isCodeRight(parameters[2])==False:
                        await message.channel.send("방 코드는 5자리의 숫자/영문으로 이루어져 있어야 합니다.")
                        return
                    elif len(parameters[3])>8:
                        await message.channel.send("비밀번호는 8자리 이하입니다.")
                        return
                    
                    try:
                        int(parameters[3])
                    except:
                        await message.channel.send("비밀번호는 정수여야 합니다.")
                        return
                    
                    
                    else:
                        for idx in range(1,max_row+1):
                            if sheet["D"+str(idx)].value==str(message.author.id):
                                sheet.delete_rows(idx)
                                max_row-=1
                                break

                        sheet["A"+str(max_row+1)].value = parameters[1]
                        sheet["B"+str(max_row+1)].value = parameters[2]
                        sheet["C"+str(max_row+1)].value = parameters[3]
                        sheet["D"+str(max_row+1)].value = str(message.author.id)
                        sheet["E"+str(max_row+1)].value = str(message.author.name)

                        file.save("room.xlsx")
                        await message.channel.send("방 등록이 완료됐습니다")
                        
                        file = openpyxl.load_workbook("mmlist.xlsx")
                        sheet = file.active
                        max_row = sheet.max_row
                        if(sheet["A1"].value==None):
                            return
                        
                        for idx in range(1,max_row+1):
                            user = await client.fetch_user(int(sheet["A"+str(idx)].value))
                            await user.create_dm()
                            channel = user.dm_channel
                            await channel.send(parameters[1]+"님의 방이 생성되었습니다. -조회")
                            
access_token=os.environ["BOT_TOKEN"]
client.run(access_token)
